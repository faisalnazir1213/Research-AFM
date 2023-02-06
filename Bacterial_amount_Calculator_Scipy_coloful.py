# -*- coding: utf-8 -*-
"""
Created on Sat Feb  4 17:54:26 2023

@author: F0321
"""

import cv2
import numpy as np
import os
import openpyxl
from scipy import ndimage

# Specify the folder containing the images
folder = "E:/Bacterial_Images/Final_Images"
result_folder="e:/Bacterial_Images/results_Scipy"
# Create a list to store the results
results = []

# Loop through all files in the folder
for filename in os.listdir(folder):
    # Load the image
    img = cv2.imread(os.path.join(folder, filename))

    # Convert the image to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    # Apply thresholding to segment the bacteria
    _, thresh = cv2.threshold(gray, 13, 255, cv2.THRESH_BINARY)

    # Find connected components in the thresholded image
    labels, num_labels = ndimage.label(thresh)

    # Create a color map for the labeled bacteria
    color_map = np.random.randint(0, 255, (num_labels + 1, 3), dtype=np.uint8)

    # Color the bacteria
    colored_bacteria = color_map[labels]

    # Combine the original image and the labeled bacteria
    labeled_img = np.zeros_like(img)
    labeled_img[labels > 0] = colored_bacteria[labels > 0]

    # Save the labeled image
    labeled_filename = filename.split(".")[0] + "_labeled.png"
    cv2.imwrite(os.path.join(result_folder, labeled_filename), labeled_img)

    # Append the results for each image to the list
    results.append([filename, num_labels])

    # Display the results
    print("{}: {} bacteria detected".format(filename, num_labels))

# Write the results to an Excel spreadsheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Bacteria Counts"
sheet["A1"] = "Filename"
sheet["B1"] = "Number of Bacteria"

for i, result in enumerate(results):
    sheet.cell(row=i + 2, column=1, value=result[0])
    sheet.cell(row=i + 2, column=2, value=result[1])

wb.save("e:/Bacterial_Images/bacteria_counts_scipy.xlsx")
